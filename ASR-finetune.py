# Standard library imports
import os
import sys
import platform
import subprocess
import threading
import time
import warnings
import re
import string
import logging
from openpyxl.styles import Font
import queue
# Third-party imports
import difflib
import pandas as pd
import tkinter as tk
import openpyxl
from tkinter import filedialog, messagebox, ttk
from pathlib import Path
import whisper
from itertools import zip_longest
from openpyxl import Workbook, load_workbook
from transformers import AutoModelForSpeechSeq2Seq, AutoProcessor, pipeline
import librosa
import numpy as np

# Suppress specific warnings
warnings.filterwarnings("ignore", message="FP16 is not supported on CPU; using FP32 instead")

import ssl
import certifi
import urllib.request

def create_ssl_context():
    context = ssl.create_default_context(cafile=certifi.where())
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE
    return context

import os
import torch
import librosa
import numpy as np

def transcribe_audio_batch(file_paths, progress_queue=None):
    # Load the model and processor from Hugging Face
    model_id = "shaunliu82714/whisper-finetuned-vocab-trained"
    device = "cuda" if torch.cuda.is_available() else "cpu"
    torch_dtype = torch.float16 if torch.cuda.is_available() else torch.float32

    model = AutoModelForSpeechSeq2Seq.from_pretrained(
        model_id, torch_dtype=torch_dtype, low_cpu_mem_usage=True
    ).to(device)

    processor = AutoProcessor.from_pretrained(model_id)

    transcriptions = {}
    start_time = time.time()
    total_files = len(file_paths)

    for i, file_path in enumerate(file_paths):
        print(f"Transcribing file {i+1}/{total_files}: {file_path}")

        # Load audio file into numpy array and resample to 16kHz
        audio, sr = librosa.load(file_path, sr=16000)
        
        # Process the audio with the feature extractor
        inputs = processor(audio, sampling_rate=16000, return_tensors="pt").to(device, torch_dtype)
        
        # Transcribe using the model
        with torch.no_grad():
            generated_ids = model.generate(inputs["input_features"])
            transcription = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]
        
        transcriptions[os.path.splitext(os.path.basename(file_path))[0]] = transcription.strip()
        
        elapsed_time = time.time() - start_time
        avg_time_per_file = elapsed_time / (i + 1)
        remaining_time = avg_time_per_file * (total_files - (i + 1))
        
        if progress_queue:
            progress_queue.put((i + 1, total_files, "Transcribing audio files...", remaining_time))
    print("Completed transcriptions:", transcriptions)
    return transcriptions

def save_transcriptions_to_excel(transcriptions, output_file='transcriptions.xlsx'):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transcriptions"
    
    # Add headers
    ws['A1'] = "VOID (File Name)"
    ws['B1'] = "Transcribed Content"
    
    # Add data
    for row, (file_name, content) in enumerate(transcriptions.items(), start=2):
        ws.cell(row=row, column=1, value=file_name)
        ws.cell(row=row, column=2, value=content)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 100
    
    # Save the workbook
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(script_dir, output_file)
    wb.save(output_path)
    print(f"Transcriptions saved to {output_path}")
    return output_path

def label_sentences(text1, text2):
    def normalize_text(text):
        return ''.join(e for e in text.lower() if e.isalnum())

    normalized_text1 = normalize_text(text1)
    normalized_text2 = normalize_text(text2)

    return normalized_text1 == normalized_text2

def filter_similar_sentences(transcriptions, show_similar=False):
    filtered_transcriptions = {}
    for key, value in transcriptions.items():
        script_text = value.get('script', '')
        transcription_text = value.get('transcription', '')
        if label_sentences(script_text, transcription_text):
            if show_similar:
                filtered_transcriptions[key] = value
        else:
            filtered_transcriptions[key] = value
    return filtered_transcriptions

def compare_texts(transcribed_text, script_text):
    transcribed_text = transcribed_text.strip()
    script_text = script_text.strip()
    
    similarity_percentage = difflib.SequenceMatcher(None, transcribed_text, script_text).ratio() * 100
    if similarity_percentage == 100:
        return similarity_percentage, transcribed_text

    diff = difflib.ndiff(script_text.split(), transcribed_text.split())
    diff_text = []
    for word in diff:
        if word.startswith('- '):
            diff_text.append(f'<del>{word[2:]}</del>')
        elif word.startswith('+ '):
            diff_text.append(f'<ins>{word[2:]}</ins>')
        else:
            diff_text.append(word[2:])
    diff_text = ' '.join(diff_text).replace('^', '')
    return similarity_percentage, diff_text

def compare_batch(transcriptions, scripts, progress_queue=None):
    comparison_results = {}
    start_time = time.time()
    total_files = len(transcriptions)
    remaining_time = 0  # Initialize remaining_time

    for i, (file_name, transcribed_text) in enumerate(transcriptions.items()):
        print(f"Comparing transcription for file: {file_name}")
        script_text = scripts.get(file_name, "")
        if script_text is None:
            print(f"No script found for {file_name}")
            continue
        similarity_percentage, delta = compare_texts(transcribed_text, script_text)
        comparison_results[file_name] = {
            "similarity_percentage": similarity_percentage,
            "differences": delta,
            "transcription": transcribed_text,
            "script": script_text
        }
        elapsed_time = time.time() - start_time
        avg_time_per_file = elapsed_time / (i + 1)
        remaining_time = avg_time_per_file * (total_files - (i + 1))
        
        if progress_queue:
            progress_queue.put((i + 1, total_files, "Comparing transcriptions with scripts...", remaining_time))
    print("Completed comparisons:", comparison_results)
    return comparison_results

def read_scripts_from_excel_files(excel_files, file_name_col, script_text_col):
    scripts = {}
    for excel_file in excel_files:
        xls = pd.ExcelFile(excel_file)
        for sheet_name in xls.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            for index, row in df.iterrows():
                file_name = row[file_name_col]
                script_text = str(row[script_text_col]).strip()
                scripts[file_name] = script_text
    print("Completed reading scripts:", scripts)
    return scripts

def main(audio_files_directory, excel_files, file_name_col, script_text_col, progress_callback=None):
    audio_files = list(Path(audio_files_directory).glob("*.wav"))
    
    transcriptions = transcribe_audio_batch(audio_files, progress_callback=progress_callback)
    
    scripts = read_scripts_from_excel_files(excel_files, file_name_col, script_text_col)
    
    comparison_results = compare_batch(transcriptions, scripts, progress_callback=progress_callback)
    
    save_transcriptions_to_excel(transcriptions)

    return comparison_results

def play_audio(file_path):
    if platform.system() == "Darwin":
        subprocess.call(('open', file_path))
    elif platform.system() == "Windows":
        os.startfile(file_path)
    else:
        subprocess.call(('xdg-open', file_path))

class ScrollableFrame(tk.Frame):
    def __init__(self, parent, *args, **kwargs):
        tk.Frame.__init__(self, parent, *args, **kwargs)
        self.canvas = tk.Canvas(self, borderwidth=0)
        self.frame = tk.Frame(self.canvas)
        self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.hsb = tk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.vsb.set, xscrollcommand=self.hsb.set)

        self.vsb.pack(side="right", fill="y")
        self.hsb.pack(side="bottom", fill="x")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas_window = self.canvas.create_window((4, 4), window=self.frame, anchor="nw", tags="self.frame")

        self.frame.bind("<Configure>", self.on_frame_configure)
        self.canvas.bind("<Configure>", self.on_canvas_configure)

        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        self.canvas.bind_all("<Button-4>", self._on_mousewheel)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel)

    def on_frame_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def on_canvas_configure(self, event):
        canvas_width = event.width
        self.canvas.itemconfig(self.canvas_window, width=canvas_width)

    def _on_mousewheel(self, event):
        if self.vsb.get() != (0.0, 1.0):
            if event.num == 4 or event.delta > 0:
                self.canvas.yview_scroll(-1, "units")
            elif event.num == 5 or event.delta < 0:
                self.canvas.yview_scroll(1, "units")

class TranscriptionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("ASR Transcription Comparison Tool")
        self.confirmation_result = False
        self.modified_scripts = {}
        self.items_per_page = 10
        self.current_page = 0
        self.total_pages = 0
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.style.configure('TButton', padding=6, relief="flat", background="#4CAF50", foreground="white")
        self.style.configure('TLabel', padding=6, font=('Helvetica', 10))
        self.style.configure('TEntry', padding=6)
        self.style.configure('TProgressbar', thickness=20, troughcolor='#F0F0F0', background='#4CAF50')

        self.setup_frame = ttk.Frame(root, padding="20 20 20 0")
        self.setup_frame.pack(fill="x")
        
        self.progress_frame = ttk.Frame(root, padding="20 10")
        self.progress_frame.pack(fill="x")
        
        self.results_frame = ScrollableFrame(root)
        self.results_frame.pack(fill="both", expand=True, padx=20, pady=10)

        self.sort_order = tk.StringVar(value="descending")
        self.results = None
        self.audio_files_directory = None
        self.show_cap_diff = tk.BooleanVar(value=True)
        self.show_punct_diff = tk.BooleanVar(value=True)
        self.show_word_diff = tk.BooleanVar(value=True)
        self.show_no_diff = tk.BooleanVar(value=True)
        self.transcribe_only_button = ttk.Button(self.setup_frame, text="Transcribe Only", command=self.run_transcribe_only)
        self.transcribe_only_button.grid(row=4, column=2, pady=10)

        self.create_setup_widgets()
        self.create_progress_widgets()
        self.create_filter_widgets()

        self.text_modifications = {}  # New dictionary to store just the text modifications

    def run_transcribe_only(self):
        audio_files_directory = self.audio_dir_entry.get()
        if not audio_files_directory:
            messagebox.showerror("Error", "Please select an audio files directory.")
            return
        
        def task():
            try:
                audio_files = list(Path(audio_files_directory).glob("*.wav"))
                transcriptions = transcribe_audio_batch(audio_files, progress_callback=self.update_progress)
                save_transcriptions_to_excel(transcriptions)
                self.root.after(0, messagebox.showinfo, "Success", "Transcriptions saved to transcriptions.xlsx")
            except Exception as e:
                self.root.after(0, messagebox.showerror, "Error", str(e))
                print(f"Error details: {e}")  
        
        threading.Thread(target=task).start()

    def close_comparison_dialog(self, window, result):
        self.confirmation_result = result
        window.destroy()
        
    def create_setup_widgets(self):
        ttk.Label(self.setup_frame, text="Audio Files Directory:").grid(row=0, column=0, sticky="w")
        self.audio_dir_entry = ttk.Entry(self.setup_frame, width=50)
        self.audio_dir_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(self.setup_frame, text="Browse", command=self.browse_audio_dir).grid(row=0, column=2, padx=5)

        ttk.Label(self.setup_frame, text="Script Excel File:").grid(row=1, column=0, sticky="w")
        self.excel_file_entry = ttk.Entry(self.setup_frame, width=50)
        self.excel_file_entry.grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(self.setup_frame, text="Browse", command=self.browse_excel_file).grid(row=1, column=2, padx=5)

        ttk.Label(self.setup_frame, text="File Name Column:").grid(row=2, column=0, sticky="w")
        self.file_name_col_combo = ttk.Combobox(self.setup_frame, state="readonly", width=47)
        self.file_name_col_combo.grid(row=2, column=1, padx=5, pady=5)

        ttk.Label(self.setup_frame, text="Script Text Column:").grid(row=3, column=0, sticky="w")
        self.script_text_col_combo = ttk.Combobox(self.setup_frame, state="readonly", width=47)
        self.script_text_col_combo.grid(row=3, column=1, padx=5, pady=5)

        self.run_button = ttk.Button(self.setup_frame, text="Run Comparison", command=self.run_tool)
        self.run_button.grid(row=4, column=1, pady=10)

    def create_progress_widgets(self):
        self.progress_label = ttk.Label(self.progress_frame, text="")
        self.progress_label.pack(side="left", padx=5)
        
        self.progress = ttk.Progressbar(self.progress_frame, length=300, mode='determinate')
        self.progress.pack(side="left", padx=5)

    def create_filter_widgets(self):
        self.filter_frame = ttk.Frame(self.root, padding="20 10")
        self.filter_frame.pack(fill="x")
        
        ttk.Label(self.filter_frame, text="Filter Results:").pack(side="left")
        ttk.Checkbutton(self.filter_frame, text="Cap. Diff", variable=self.show_cap_diff).pack(side="left")
        ttk.Checkbutton(self.filter_frame, text="Punct. Diff", variable=self.show_punct_diff).pack(side="left")
        ttk.Checkbutton(self.filter_frame, text="Word Diff", variable=self.show_word_diff).pack(side="left")
        ttk.Checkbutton(self.filter_frame, text="No Diff", variable=self.show_no_diff).pack(side="left")
        
        ttk.Label(self.filter_frame, text="Sort by:").pack(side="left", padx=(20, 0))
        ttk.Radiobutton(self.filter_frame, text="Descending", variable=self.sort_order, value="descending").pack(side="left")
        ttk.Radiobutton(self.filter_frame, text="Ascending", variable=self.sort_order, value="ascending").pack(side="left")
        
        self.apply_button = ttk.Button(self.filter_frame, text="Apply Filters", command=self.apply_filters_and_sort)
        self.apply_button.pack(side="left", padx=10)
        
        # Add Save All Changes button
        self.save_button = ttk.Button(self.filter_frame, text="Save All Changes", command=self.save_all_modified_scripts)
        self.save_button.pack(side="left", padx=10)
        
        self.filter_frame.pack_forget()  # Hide initially

    def apply_filters_and_sort(self):
        if self.all_results and self.audio_files_directory:
            filtered_results = self.filter_results(self.all_results)
            sorted_results = self.sort_results_by_similarity(filtered_results)
            self.all_results = sorted_results
            self.total_pages = (len(sorted_results) - 1) // self.items_per_page + 1
            self.current_page = 0
            self.display_current_page()

    def filter_results(self, results):
        filtered = {}
        for file_name, result in results.items():
            diff_types = self.check_differences(result['script'], result['transcription'])
            if any((diff_type == "capitalization" and self.show_cap_diff.get()) or
                   (diff_type == "punctuation" and self.show_punct_diff.get()) or
                   (diff_type == "word" and self.show_word_diff.get()) or
                   (diff_type == "none" and self.show_no_diff.get())
                   for diff_type in diff_types):
                filtered[file_name] = result
        return filtered

    def check_differences(self, script, transcription):
        differences = set()
        
        # Check for word differences
        script_words = re.findall(r'\b\w+\b', script.lower())
        trans_words = re.findall(r'\b\w+\b', transcription.lower())
        if script_words != trans_words:
            differences.add("word")
        
        # Check for punctuation differences
        script_punct = re.findall(r'[^\w\s]', script)
        trans_punct = re.findall(r'[^\w\s]', transcription)
        if script_punct != trans_punct:
            differences.add("punctuation")
        
        # Check for capitalization differences
        if script.split() != transcription.split() and script.lower().split() == transcription.lower().split():
            differences.add("capitalization")
        
        return list(differences) if differences else ["none"]

    def highlight_differences(self, text_widget, script, transcription):
        text_widget.tag_configure('word_diff', foreground='red', underline=True)
        text_widget.tag_configure('punct_diff', foreground='purple', underline=True)
        text_widget.tag_configure('cap_diff', foreground='blue', underline=True)
        text_widget.tag_configure('normal', foreground='black')

        text_widget.delete("1.0", tk.END)
        
        s_words = script.split()
        t_words = transcription.split()
        
        matcher = difflib.SequenceMatcher(None, s_words, t_words)
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                text_widget.insert(tk.END, ' '.join(t_words[j1:j2]) + ' ', 'normal')
            elif tag in ('replace', 'insert'):
                for word in t_words[j1:j2]:
                    if word.lower() in [w.lower() for w in s_words[i1:i2]]:
                        text_widget.insert(tk.END, word + ' ', 'cap_diff')
                    elif any(c in word for c in string.punctuation):
                        text_widget.insert(tk.END, word + ' ', 'punct_diff')
                    else:
                        text_widget.insert(tk.END, word + ' ', 'word_diff')

        # Remove trailing space
        text_widget.delete("end-2c", tk.END)

    def get_diff_color(self, diff_types):
        color_priority = {"word": "red", "punctuation": "purple", "capitalization": "blue", "none": "green"}
        for diff_type in color_priority:
            if diff_type in diff_types:
                return color_priority[diff_type]
        return "black"

    def show_results(self, results, audio_files_directory):
        self.modified_scripts = {}
        self.all_results = results
        self.audio_files_directory = audio_files_directory
        
        # Calculate total pages
        self.total_pages = (len(results) - 1) // self.items_per_page + 1
        self.current_page = 0
        
        # Reset pagination frame
        if hasattr(self, 'pagination_frame'):
            self.pagination_frame.destroy()
            delattr(self, 'pagination_frame')
        
        self.display_current_page()

    def display_current_page(self):
        text_width = 50

        # Clear existing content
        for widget in self.results_frame.frame.winfo_children():
            widget.destroy()

        # Create pagination frame at the top
        self.pagination_frame = ttk.Frame(self.results_frame.frame)
        self.pagination_frame.pack(side="top", fill="x", pady=5)
        
        # Previous page button
        self.prev_button = ttk.Button(
            self.pagination_frame, 
            text="Previous", 
            command=self.prev_page,
            state="disabled" if self.current_page == 0 else "normal"
        )
        self.prev_button.pack(side="left", padx=5)
        
        # Page indicator label
        self.page_label = ttk.Label(
            self.pagination_frame, 
            text=f"Page {self.current_page + 1} of {self.total_pages}"
        )
        self.page_label.pack(side="left", padx=5)
        
        # Next page button
        self.next_button = ttk.Button(
            self.pagination_frame, 
            text="Next", 
            command=self.next_page,
            state="disabled" if self.current_page >= self.total_pages - 1 else "normal"
        )
        self.next_button.pack(side="left", padx=5)

        # Display current page content
        start_index = self.current_page * self.items_per_page
        end_index = start_index + self.items_per_page
        current_results = dict(list(self.all_results.items())[start_index:end_index])

        for file_name, result in current_results.items():
            # Create frame for each result
            result_frame = ttk.Frame(self.results_frame.frame)
            result_frame.pack(fill="x", expand=True, pady=10)
            
            header_frame = tk.Frame(result_frame)
            header_frame.pack(fill="x")

            tk.Label(header_frame, text=f"File: {file_name}", font=('Helvetica', 12, 'bold')).pack(side="left")
            tk.Label(header_frame, text=f"Similarity: {result['similarity_percentage']:.2f}%").pack(side="left", padx=20)
            
            diff_types = self.check_differences(result['script'], result['transcription'])
            diff_labels = [diff_type.capitalize() for diff_type in diff_types if diff_type != "none"]
            diff_text = " & ".join(diff_labels) + " Diff" if diff_labels else "No Diff"
            diff_color = self.get_diff_color(diff_types)
            diff_label = tk.Label(header_frame, text=diff_text, fg=diff_color)
            diff_label.pack(side="left", padx=5)
            
            tk.Button(header_frame, text="Play Audio", command=lambda fn=file_name: play_audio(os.path.join(self.audio_files_directory, fn + '.wav'))).pack(side="right")

            comparison_frame = tk.Frame(result_frame)
            comparison_frame.pack(fill="x", expand=True)

            comparison_frame.grid_columnconfigure(0, weight=1)
            comparison_frame.grid_columnconfigure(1, weight=1)

            script_text = tk.Text(comparison_frame, height=5, wrap='word', width=text_width, bg="white", fg="black", insertbackground="black")
            
            # Check if there's a saved modification
            if file_name in self.text_modifications:
                script_text.insert(tk.END, self.text_modifications[file_name])
            else:
                script_text.insert(tk.END, result['script'])
                
            script_text.grid(row=1, column=0, padx=(0, 5), pady=5, sticky="nsew")

            def on_text_change(event, fn=file_name):
                current_text = event.widget.get("1.0", "end-1c")
                if current_text != result['script']:
                    self.text_modifications[fn] = current_text
                    self.modified_scripts[fn] = {
                        'original': result['script'],
                        'modified': current_text,
                        'transcription': result['transcription']
                    }
                elif fn in self.text_modifications:
                    del self.text_modifications[fn]
                    if fn in self.modified_scripts:
                        del self.modified_scripts[fn]

            script_text.bind('<KeyRelease>', on_text_change)

            asr_text = tk.Text(comparison_frame, height=5, wrap='word', width=text_width, bg="white", fg="black")
            self.highlight_differences(asr_text, result['script'], result['transcription'])
            asr_text.config(state=tk.DISABLED)
            asr_text.grid(row=1, column=1, padx=(5, 0), pady=5, sticky="nsew")

            script_scrollbar = tk.Scrollbar(comparison_frame, command=script_text.yview)
            script_scrollbar.grid(row=1, column=0, sticky='nse')
            script_text.config(yscrollcommand=script_scrollbar.set)

            asr_scrollbar = tk.Scrollbar(comparison_frame, command=asr_text.yview)
            asr_scrollbar.grid(row=1, column=1, sticky='nse')
            asr_text.config(yscrollcommand=asr_scrollbar.set)

            # Enable editing by default and ensure cursor visibility
            script_text.config(state=tk.NORMAL, insertbackground="black")

        # Make sure the pagination frame stays on top
        if hasattr(self, 'pagination_frame'):
            self.pagination_frame.lift()

    def prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.display_current_page()

    def next_page(self):
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.display_current_page()

    # Set up logging
    logging.basicConfig(level=logging.DEBUG, filename='script_debug.log', filemode='w',
                        format='%(asctime)s - %(levelname)s - %(message)s')
    import openpyxl
    def save_all_modified_scripts(self):
        print("Starting save_all_modified_scripts function")
        print(f"Number of modifications tracked: {len(self.text_modifications)}")  # Debug print
        
        if not self.text_modifications:
            messagebox.showinfo("No Changes", "No modifications were detected. No file will be saved.")
            return

        # Create a new workbook for modified scripts
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Modified Scripts"

        # Add headers
        headers = ["VOID", "Original Line", "ASR Result", "Modified Line"]
        for col, header in enumerate(headers, start=1):
            new_sheet.cell(row=1, column=col, value=header)
            new_sheet.cell(row=1, column=col).font = Font(bold=True)

        modifications = {}
        for file_name, modified_text in self.text_modifications.items():
            original_script = self.all_results[file_name]['script']
            asr_result = self.all_results[file_name]['transcription']
            
            modifications[file_name] = {
                'original': original_script,
                'modified': modified_text,
                'asr': asr_result
            }

        if modifications:
            if self.show_all_comparisons(modifications):
                row = 2
                for file_name, data in modifications.items():
                    new_sheet.cell(row=row, column=1, value=file_name)
                    new_sheet.cell(row=row, column=2, value=data['original'])
                    new_sheet.cell(row=row, column=3, value=data['asr'])
                    new_sheet.cell(row=row, column=4, value=data['modified'])
                    row += 1

                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
                )
                if save_path:
                    new_workbook.save(save_path)
                    messagebox.showinfo("Success", 
                        f"Modified scripts have been saved. Total modifications: {len(modifications)}")
        else:
            messagebox.showinfo("No Changes", "No modifications were detected. No file will be saved.")

    def show_all_comparisons(self, modifications):
        # Create a new window for comparison
        comparison_window = tk.Toplevel(self.root)
        comparison_window.title("Review All Modifications")

        # Create a scrollable frame
        canvas = tk.Canvas(comparison_window)
        scrollbar = tk.Scrollbar(comparison_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Add headers
        tk.Label(scrollable_frame, text="File Name", font=('Helvetica', 12, 'bold')).grid(row=0, column=0, padx=5, pady=5)
        tk.Label(scrollable_frame, text="Original Script", font=('Helvetica', 12, 'bold')).grid(row=0, column=1, padx=5, pady=5)
        tk.Label(scrollable_frame, text="Modified Script", font=('Helvetica', 12, 'bold')).grid(row=0, column=2, padx=5, pady=5)
        tk.Label(scrollable_frame, text="ASR Result", font=('Helvetica', 12, 'bold')).grid(row=0, column=3, padx=5, pady=5)

        # Function to highlight differences
        def highlight_differences(text_widget, base_text, compare_text):
            text_widget.config(state=tk.NORMAL)
            text_widget.delete("1.0", tk.END)
            matcher = difflib.SequenceMatcher(None, base_text.split(), compare_text.split())
            for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                if tag == 'equal':
                    text_widget.insert(tk.END, ' '.join(base_text.split()[i1:i2]) + ' ', 'normal')
                else:
                    text_widget.insert(tk.END, ' '.join(compare_text.split()[j1:j2]) + ' ', 'diff')
            text_widget.config(state=tk.DISABLED)

        # Populate the scrollable frame with modifications
        for i, (file_name, data) in enumerate(modifications.items(), start=1):
            original_script = data['original']
            modified_script = data['modified']
            asr_result = data['asr']

            tk.Label(scrollable_frame, text=file_name).grid(row=i, column=0, padx=5, pady=5)

            original_text = tk.Text(scrollable_frame, height=5, width=30, wrap='word', bg="lightgrey", fg="black")
            original_text.insert(tk.END, original_script)
            original_text.config(state=tk.DISABLED)
            original_text.grid(row=i, column=1, padx=5, pady=5, sticky="nsew")

            modified_text = tk.Text(scrollable_frame, height=5, width=30, wrap='word', bg="lightgrey", fg="black")
            modified_text.tag_configure('diff', background='yellow', foreground='black')
            modified_text.tag_configure('normal', background='white', foreground='black')
            highlight_differences(modified_text, original_script, modified_script)
            modified_text.grid(row=i, column=2, padx=5, pady=5, sticky="nsew")

            asr_text = tk.Text(scrollable_frame, height=5, width=30, wrap='word', bg="lightgrey", fg="black")
            asr_text.insert(tk.END, asr_result)
            asr_text.config(state=tk.DISABLED)
            asr_text.grid(row=i, column=3, padx=5, pady=5, sticky="nsew")

        # Pack the canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Add a global confirm button
        confirm_button = tk.Button(comparison_window, text="Confirm All", command=lambda: self.close_comparison_dialog(comparison_window, True))
        confirm_button.pack(pady=10)

        # Wait for user action
        self.root.wait_window(comparison_window)
        return self.confirmation_result

    def sort_results_by_similarity(self, results):
        sorted_results = sorted(
            results.items(),
            key=lambda x: x[1]['similarity_percentage'],
            reverse=(self.sort_order.get() == "descending")
        )
        return dict(sorted_results)

    def browse_audio_dir(self):
        directory = filedialog.askdirectory()
        self.audio_dir_entry.delete(0, tk.END)
        self.audio_dir_entry.insert(0, directory)

    def browse_excel_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        self.excel_file_entry.delete(0, tk.END)
        self.excel_file_entry.insert(0, file_path)
        self.update_column_options(file_path)

    def update_column_options(self, excel_file_path):
        try:
            df = pd.read_excel(excel_file_path, nrows=0)
            columns = list(df.columns)
            self.file_name_col_combo['values'] = columns
            self.script_text_col_combo['values'] = columns
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read columns from Excel file: {str(e)}")

    def update_progress(self, current, total, status="", remaining_time=0):
        def update():
            self.progress["value"] = (current / total) * 100
            self.progress_label.config(text=f"{status} ({current}/{total}) - Est. time: {int(remaining_time // 60)}m {int(remaining_time % 60)}s")
            self.root.update_idletasks()

        self.root.after(0, update)

    def run_tool(self):
        audio_files_directory = self.audio_dir_entry.get()
        excel_file_path = self.excel_file_entry.get()
        file_name_col = self.file_name_col_combo.get()
        script_text_col = self.script_text_col_combo.get()
        if not all([audio_files_directory, excel_file_path, file_name_col, script_text_col]):
            messagebox.showerror("Error", "Please fill in all fields before running the comparison.")
            return

        self.progress["value"] = 0
        self.progress_label.config(text="")

        progress_queue = queue.Queue()

        def task():
            try:
                # Transcription phase
                audio_files = list(Path(audio_files_directory).glob("*.wav"))
                transcriptions = transcribe_audio_batch(audio_files, progress_queue)

                # Read scripts
                scripts = read_scripts_from_excel_files([excel_file_path], file_name_col, script_text_col)

                # Comparison phase
                results = compare_batch(transcriptions, scripts, progress_queue)

                self.all_results = results
                self.audio_files_directory = audio_files_directory
                self.root.after(0, self.show_results, results, audio_files_directory)
                self.root.after(0, self.filter_frame.pack)
            except Exception as e:
                self.root.after(0, messagebox.showerror, "Error", str(e))

        def update_gui():
            try:
                while True:
                    current, total, status, remaining_time = progress_queue.get_nowait()
                    self.update_progress(current, total, status, remaining_time)
            except queue.Empty:
                pass
            finally:
                self.root.after(100, update_gui)

        threading.Thread(target=task, daemon=True).start()
        self.root.after(100, update_gui)

if __name__ == "__main__":
    root = tk.Tk()
    app = TranscriptionApp(root)
    root.mainloop()
